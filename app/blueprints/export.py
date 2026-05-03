"""
Export Blueprint – Excel ve PDF dışa aktarım (SRP).
"""

from flask import Blueprint, request, send_file
from datetime import datetime, date as date_type
import io

from db_manager import get_db_connection
from app.decorators import admin_required, hoca_required
from app.utils import get_admin_bolum_ids

export_bp = Blueprint('export', __name__)


def _get_export_rows(tur, bolum_id=None, bolum_ids=None):
    """
    Panel ile birebir aynı slot mantığını uygular.
    Python tarafında dedup: aynı hafta+konu için onaylı slot tercih edilir.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    if bolum_id:
        cursor.execute("""
            SELECT sp.sunumid, sp.haftano, k.konuadi, sp.sunumtarihi, k.sirano
            FROM sunumprogrami sp
            JOIN konular k ON k.konuid = sp.konuid
            WHERE sp.bolumid = %s
               OR (sp.bolumid IS NULL AND sp.ogretimturu = (
                      SELECT ogretimturu FROM bolumler WHERE bolumid = %s LIMIT 1))
            ORDER BY sp.haftano, k.sirano, sp.bolumid ASC NULLS LAST
        """, (bolum_id, bolum_id))
    elif bolum_ids:
        cursor.execute("""
            SELECT sp.sunumid, sp.haftano, k.konuadi, sp.sunumtarihi, k.sirano
            FROM sunumprogrami sp
            JOIN konular k ON k.konuid = sp.konuid
            WHERE sp.bolumid = ANY(%s)
               OR (sp.bolumid IS NULL AND sp.ogretimturu = %s)
            ORDER BY sp.haftano, k.sirano, sp.bolumid ASC NULLS LAST
        """, (bolum_ids, tur))
    else:
        cursor.execute("""
            SELECT sp.sunumid, sp.haftano, k.konuadi, sp.sunumtarihi, k.sirano
            FROM sunumprogrami sp
            JOIN konular k ON k.konuid = sp.konuid
            WHERE sp.ogretimturu = %s
            ORDER BY sp.haftano, k.sirano, sp.bolumid ASC NULLS LAST
        """, (tur,))

    slots = cursor.fetchall()
    if not slots:
        return []

    slot_ids = [s['sunumid'] for s in slots]

    # Onaylı öğrencileri topluca çek; isim boşsa öğrenci numarasıyla göster
    cursor.execute("""
        SELECT sg.sunumid,
               COALESCE(NULLIF(TRIM(o.adsoyad), ''), o.ogrencino, '(Bilinmiyor)') AS display_name
        FROM sunumgorevlileri sg
        JOIN ogrenciler o ON o.ogrenciid = sg.ogrenciid
        WHERE sg.sunumid = ANY(%s)
        ORDER BY sg.gorevid
    """, (slot_ids,))
    atananlar_map = {}
    for r in cursor.fetchall():
        atananlar_map.setdefault(r['sunumid'], []).append(r['display_name'])

    # Python-side dedup: (haftano, konuadi) başına en iyi slotu tut.
    # Onaylı slot yoksa, onaylı olanı tercih et.
    seen = {}   # key → {'s': slot_row, 'atananlar': list}
    for s in slots:
        key        = (s['haftano'], s['konuadi'])
        atananlar  = atananlar_map.get(s['sunumid'], [])
        if key not in seen:
            seen[key] = {'s': s, 'atananlar': atananlar}
        elif atananlar and not seen[key]['atananlar']:
            # Mevcut slot onaylanmamış; bu onaylı → güncelle
            seen[key] = {'s': s, 'atananlar': atananlar}

    result = []
    for v in seen.values():
        s         = v['s']
        atananlar = v['atananlar']
        result.append({
            'sunumid'   : s['sunumid'],
            'haftano'   : s['haftano'],
            'konuadi'   : s['konuadi'],
            'sunumtarihi': s['sunumtarihi'],
            'sirano'    : s['sirano'],
            'atananlar' : ' & '.join(a for a in atananlar if a) if any(atananlar) else '—',
            'onaylandi' : bool(atananlar),
        })

    result.sort(key=lambda x: (
        x['sunumtarihi'] or date_type.max,
        x['haftano'],
        x['sirano'],
    ))
    return result


def _resolve_bolum_ids(tur, bolum_id=None, bolum_ids=None):
    """Export için ilgili bölüm ID listesini döndürür."""
    if bolum_id:
        return [int(bolum_id)]
    if bolum_ids:
        return list(bolum_ids)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT bolumid FROM bolumler WHERE ogretimturu = %s", (tur,))
    ids = [r['bolumid'] for r in cursor.fetchall()]
    return ids


def _get_special_weeks(bolum_ids_list):
    """Vize ve tatil kaydirma haftalarini döndürür: {haftano: {hafta_tarihi, aciklama}}"""
    if not bolum_ids_list:
        return {}, {}
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT haftano, MIN(hafta_tarihi) AS hafta_tarihi,
               string_agg(DISTINCT COALESCE(aciklama, 'Vize'), ', ') AS aciklama
        FROM vizehaftalari WHERE bolumid = ANY(%s)
        GROUP BY haftano
    """, (bolum_ids_list,))
    vize_map = {r['haftano']: {'hafta_tarihi': r['hafta_tarihi'], 'aciklama': r['aciklama']}
                for r in cursor.fetchall()}

    cursor.execute("""
        SELECT haftano, MIN(hafta_tarihi) AS hafta_tarihi,
               string_agg(DISTINCT COALESCE(aciklama, 'Tatil'), ', ') AS aciklama
        FROM tatilkaydirmahaftalari WHERE bolumid = ANY(%s)
        GROUP BY haftano
    """, (bolum_ids_list,))
    tatil_map = {r['haftano']: {'hafta_tarihi': r['hafta_tarihi'], 'aciklama': r['aciklama']}
                 for r in cursor.fetchall()}

    return vize_map, tatil_map


def _build_excel(tur, bolum_id=None, bolum_ids=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    rows = _get_export_rows(tur, bolum_id, bolum_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sunum Programı"

    header_fill = PatternFill("solid", fgColor="1e293b")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color="d1d5db")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Hafta No", "Konu Adı", "Atanan Öğrenci(ler)", "Durum"]
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 14

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    green_fill = PatternFill("solid", fgColor="d1fae5")
    for i, row in enumerate(rows, 2):
        atanan = row['atananlar']
        durum = "Onaylandı" if atanan != "—" else "Bekliyor"
        values = [row['haftano'], row['konuadi'], atanan, durum]
        fill = green_fill if atanan != "—" else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_pdf(tur, bolum_id=None, bolum_ids=None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # ── Font kaydı (Türkçe karakter desteği) ────────────────────────────────
    # Windows'ta Arial, Docker/Linux'ta DejaVu (ReportLab ile gelir)
    _font_candidates = [
        ("C:/Windows/Fonts/arial.ttf",   "C:/Windows/Fonts/arialbd.ttf"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ]
    fn, fnb = "Helvetica", "Helvetica-Bold"   # ultimate fallback
    for _fp, _fpb in _font_candidates:
        if os.path.exists(_fp):
            try:
                pdfmetrics.registerFont(TTFont("ExportFont",     _fp))
                pdfmetrics.registerFont(TTFont("ExportFont-Bold", _fpb if os.path.exists(_fpb) else _fp))
                fn, fnb = "ExportFont", "ExportFont-Bold"
            except Exception:
                pass
            break

    # ── Renk paleti ──────────────────────────────────────────────────────────
    NAVY     = colors.HexColor('#0f172a')
    TEAL     = colors.HexColor('#0d9488')
    TEAL_LT  = colors.HexColor('#e0f7f4')
    GREEN    = colors.HexColor('#15803d')
    GREEN_BG = colors.HexColor('#dcfce7')
    ORANGE   = colors.HexColor('#c2410c')
    GRAY     = colors.HexColor('#f8fafc')
    BORDER   = colors.HexColor('#cbd5e1')
    WHITE    = colors.white
    TEXT     = colors.HexColor('#1e293b')

    # ── Stil tanımları ───────────────────────────────────────────────────────
    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    title_st = ps('ti',  fontName=fnb, fontSize=15, textColor=NAVY,   leading=19)
    info_st  = ps('in',  fontName=fn,  fontSize=8,  textColor=colors.HexColor('#475569'), leading=12)
    wk_st    = ps('wk',  fontName=fnb, fontSize=10, textColor=WHITE,  leading=14)
    ch_st    = ps('ch',  fontName=fnb, fontSize=8,  textColor=WHITE,  leading=11)
    cell_st  = ps('ce',  fontName=fn,  fontSize=8,  textColor=TEXT,   leading=11)
    cell_ok  = ps('co',  fontName=fnb, fontSize=8,  textColor=GREEN,  leading=11)
    ok_st    = ps('ok',  fontName=fnb, fontSize=8,  textColor=GREEN,  alignment=1, leading=11)
    wait_st  = ps('wt',  fontName=fn,  fontSize=8,  textColor=ORANGE, alignment=1, leading=11)
    num_st   = ps('nu',  fontName=fn,  fontSize=8,  textColor=TEXT,   alignment=1, leading=11)

    rows = _get_export_rows(tur, bolum_id, bolum_ids)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elems = []

    # ── Başlık banner'ı ──────────────────────────────────────────────────────
    onay_c = sum(1 for r in rows if r['onaylandi'])
    bekl_c = len(rows) - onay_c
    banner = Table([[
        Paragraph(f"SUNUM PROGRAMI  —  {tur.upper()}", title_st),
        Paragraph(
            f"Oluşturma: {datetime.now().strftime('%d.%m.%Y  %H:%M')}<br/>"
            f"<font color='#15803d'>● {onay_c} Onaylı</font>"
            f"  <font color='#c2410c'>○ {bekl_c} Bekliyor</font>",
            info_st),
    ]], colWidths=[20.7*cm, 5.5*cm])
    banner.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), TEAL_LT),
        ('BOX',           (0,0), (-1,-1), 1.5, TEAL),
        ('LEFTPADDING',   (0,0), (-1,-1), 14),
        ('RIGHTPADDING',  (0,0), (-1,-1), 14),
        ('TOPPADDING',    (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',         (1,0), (1,-1),  'RIGHT'),
    ]))
    elems.append(banner)
    elems.append(Spacer(1, 0.6*cm))

    # ── Hafta bazlı tablolar ─────────────────────────────────────────────────
    COL_W  = [1.8*cm, 9*cm, 12.5*cm, 3.4*cm]
    CH_ROW = [Paragraph('Hafta', ch_st), Paragraph('Konu Adı', ch_st),
              Paragraph('Atanan Öğrenci(ler)', ch_st), Paragraph('Durum', ch_st)]

    def flush(week_label, data):
        """data: list of (is_approved, [cell0, cell1, cell2, cell3])"""
        if not data or not week_label:
            return
        t_rows = [[Paragraph(week_label, wk_st), '', '', ''], CH_ROW]
        approved_idxs, odd_idxs = [], []
        for i, (approved, cells) in enumerate(data):
            t_rows.append(cells)
            ri = i + 2
            if approved:
                approved_idxs.append(ri)
            elif ri % 2 == 0:
                odd_idxs.append(ri)

        cmds = [
            ('SPAN',          (0,0), (-1,0)),
            ('BACKGROUND',    (0,0), (-1,0),  TEAL),
            ('TOPPADDING',    (0,0), (-1,0),  9),
            ('BOTTOMPADDING', (0,0), (-1,0),  9),
            ('LEFTPADDING',   (0,0), (-1,0),  12),
            ('BACKGROUND',    (0,1), (-1,1),  NAVY),
            ('TOPPADDING',    (0,1), (-1,1),  6),
            ('BOTTOMPADDING', (0,1), (-1,1),  6),
            ('GRID',          (0,0), (-1,-1), 0.35, BORDER),
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING',   (0,2), (-1,-1), 8),
            ('RIGHTPADDING',  (0,2), (-1,-1), 8),
            ('TOPPADDING',    (0,2), (-1,-1), 5),
            ('BOTTOMPADDING', (0,2), (-1,-1), 5),
            ('ALIGN',         (0,2), (0,-1),  'CENTER'),
            ('ALIGN',         (3,2), (3,-1),  'CENTER'),
        ]
        for ri in approved_idxs:
            cmds.append(('BACKGROUND', (0, ri), (-1, ri), GREEN_BG))
        for ri in odd_idxs:
            cmds.append(('BACKGROUND', (0, ri), (-1, ri), GRAY))

        t = Table(t_rows, colWidths=COL_W, repeatRows=2)
        t.setStyle(TableStyle(cmds))
        elems.append(t)
        elems.append(Spacer(1, 0.45*cm))

    # ── Vize / Tatil özel hafta stilleri ──────────────────────────────────────
    PURPLE    = colors.HexColor('#6d28d9')
    PURPLE_LT = colors.HexColor('#ede9fe')
    ORNG      = colors.HexColor('#ea580c')
    ORNG_LT   = colors.HexColor('#fff7ed')

    sp_st = ps('sp', fontName=fnb, fontSize=9, textColor=WHITE, leading=13)

    def add_special_week(label_text, bg, border):
        t = Table([[Paragraph(label_text, sp_st), '', '', '']],
                  colWidths=COL_W)
        t.setStyle(TableStyle([
            ('SPAN',          (0,0), (-1,-1)),
            ('BACKGROUND',    (0,0), (-1,-1), bg),
            ('BOX',           (0,0), (-1,-1), 1.5, border),
            ('LEFTPADDING',   (0,0), (-1,-1), 14),
            ('TOPPADDING',    (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        elems.append(t)
        elems.append(Spacer(1, 0.45*cm))

    # ── Tüm hafta etkinliklerini birleştir ───────────────────────────────────
    bid_list = _resolve_bolum_ids(tur, bolum_id, bolum_ids)
    vize_map, tatil_map = _get_special_weeks(bid_list)

    # Tatil ve vize haftalarını hariç tut — bunlar her zaman banner olarak gösterilir
    special_weeks = set(vize_map.keys()) | set(tatil_map.keys())

    week_groups = {}
    for row in rows:
        if row['haftano'] in special_weeks:
            continue   # Bu haftanın verisi banner'ın altında gösterilmez
        k = (row['haftano'], row['sunumtarihi'])
        week_groups.setdefault(k, []).append(row)

    all_events = []

    for (hno, dt), wrows in week_groups.items():
        all_events.append({'type': 'data', 'haftano': hno, 'sunumtarihi': dt, 'rows': wrows})

    for hno, info in vize_map.items():
        all_events.append({'type': 'vize', 'haftano': hno,
                           'sunumtarihi': info['hafta_tarihi'], 'aciklama': info['aciklama']})

    for hno, info in tatil_map.items():
        all_events.append({'type': 'tatil', 'haftano': hno,
                           'sunumtarihi': info['hafta_tarihi'], 'aciklama': info['aciklama']})

    # Önce haftano'ya göre sırala; tarihi olan haftalar tarihe göre, olmayanlar hafta no sırasında
    all_events.sort(key=lambda x: (x['haftano'], x['sunumtarihi'] or date_type.max))

    for ev in all_events:
        hno = ev['haftano']
        dt  = ev['sunumtarihi']
        ds  = dt.strftime('%d.%m.%Y') if dt else 'Tarih Belirlenmedi'

        if ev['type'] == 'vize':
            add_special_week(
                f"{hno}. HAFTA  ·  {ds}  —  VIZE HAFTASI  ({ev['aciklama']})",
                PURPLE, PURPLE)

        elif ev['type'] == 'tatil':
            add_special_week(
                f"{hno}. HAFTA  ·  {ds}  —  TATIL / KAYDIRMA HAFTASI  ({ev['aciklama']})",
                ORNG, ORNG)

        else:
            label = f"{hno}. HAFTA  ·  {ds}"
            week_data = []
            for row in ev['rows']:
                approved = row['onaylandi']
                cells = [
                    Paragraph(str(row['haftano']), num_st),
                    Paragraph(row['konuadi'],      cell_st),
                    Paragraph(row['atananlar'],    cell_ok if approved else cell_st),
                    Paragraph('✓ Onaylandı'       if approved else '○ Bekliyor',
                              ok_st               if approved else wait_st),
                ]
                week_data.append((approved, cells))
            flush(label, week_data)

    doc.build(elems)
    buf.seek(0)
    return buf


# --- Admin Export ---

@export_bp.route('/admin/export/excel')
@admin_required
def admin_export_excel():
    tur = request.args.get('tur', 'Örgün')
    admin_bolum_ids = get_admin_bolum_ids()
    if admin_bolum_ids is not None and len(admin_bolum_ids) > 0:
        buf = _build_excel(tur, bolum_ids=admin_bolum_ids)
    else:
        buf = _build_excel(tur)
    fname = f"sunum_programi_{tur}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@export_bp.route('/admin/export/pdf')
@admin_required
def admin_export_pdf():
    tur = request.args.get('tur', 'Örgün')
    admin_bolum_ids = get_admin_bolum_ids()
    if admin_bolum_ids is not None and len(admin_bolum_ids) > 0:
        buf = _build_pdf(tur, bolum_ids=admin_bolum_ids)
    else:
        buf = _build_pdf(tur)
    fname = f"sunum_programi_{tur}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')


# --- Hoca Export ---

@export_bp.route('/hoca/export/excel')
@hoca_required
def hoca_export_excel():
    bolum_id = request.args.get('bolum_id', '')
    tur = request.args.get('tur', 'Örgün')
    buf = _build_excel(tur, bolum_id=bolum_id or None)
    fname = f"sunum_programi_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@export_bp.route('/hoca/export/pdf')
@hoca_required
def hoca_export_pdf():
    bolum_id = request.args.get('bolum_id', '')
    tur = request.args.get('tur', 'Örgün')
    buf = _build_pdf(tur, bolum_id=bolum_id or None)
    fname = f"sunum_programi_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')
